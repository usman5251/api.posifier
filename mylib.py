import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import platform
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import csv
import random

def GetBrowser(**kwargs):
    headless = kwargs.get('headless', False)
    disableSandbox = kwargs.get('sandbox', False)
    startMaximized = kwargs.get('start-maximized', True)
    customDownloadDirectory = kwargs.get('downloadDirectory', False)
    chrome_options = Options()
    if headless:
        chrome_options.add_argument('--headless')
    if disableSandbox:
        chrome_options.add_argument('--no-sandbox')
    if startMaximized:
        chrome_options.add_argument('--start-maximized')
    if customDownloadDirectory != False:
        prefs = {"download.default_directory" : customDownloadDirectory}
        chrome_options.add_experimental_option("prefs",prefs)

    if platform.system() == 'Windows':
        browser = webdriver.Chrome(chrome_options=chrome_options, executable_path=os.path.join(os.getcwd(), 'chromedriver.exe'))
        return browser
    elif platform.system() == 'Linux':
        browser = webdriver.Chrome(chrome_options=chrome_options, executable_path=os.path.join(os.getcwd(), 'chromedriver'))
        return browser

def GetFileList(path, FileExtension):
    sourcePath = path
    filelist = []
    for file in os.listdir(path=sourcePath):
        if file.endswith('{}'.format(FileExtension)):
            filelist.append(file)
        else:
            pass
    return filelist

def DeleteFolderContents(path, FileExtension):
    filelist = GetFileList(path, FileExtension)
    sourcePath = path
    for f in filelist:
        if platform.system() == 'Windows':
            os.system('del /f {}'.format(os.path.join(sourcePath, f)))
        elif platform.system() == 'Linux':
            os.system('rm {}'.format(os.path.join(sourcePath, f)))
        print('{} Deleted!'.format(f))
    print("ALL FILES DELETED")

def AppendCSVs(DestinationPath, DataFrame):
    print("Appending to {}".format(DestinationPath))
    df = DataFrame
    df.to_csv(DestinationPath, mode='a', header=False, index=False)

def WriteCSV(DestinationPath, DataFrame):
    print("Appending to {}".format(DestinationPath))
    df = DataFrame
    df.to_csv(DestinationPath, mode='w', header=False, index=False)

def ReadCSV(path):
    try:
        with open(path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            data = list(reader)
        return data
    except:
        with open(path, newline='', encoding='cp1252') as f:
            reader = csv.reader(f)
            data = list(reader)
        return data

def ReadXLSX(path,worksheet_number):
    XLSX = []
    wb = load_workbook(path, data_only=True)
    sh = wb.worksheets[worksheet_number]
    noRows = sh.max_row
    noCol = sh.max_column+1
    for row in range(1,noRows):
        XLSX.append([])
        for col in range(1,noCol):
            XLSX[row-1].append(sh.cell(row,col).value)
    return XLSX

def WriteXLSX(path,dataFrame):
    writer = pd.ExcelWriter(path, engine='openpyxl') # pylint: disable=abstract-class-instantiated
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    df = dataFrame
    df.to_excel(writer,index=False,header=False)
    writer.close()

def ConvertToDataFrame(Data, Columns):
    df = pd.DataFrame(Data)
    # df = df.transpose()
    # df.columns = Columns
    return df

def WriteToSpecificColCSV(path, r, c, value):
    with open(path) as f:
        reader = csv.reader(f)
        data = list(reader)
        try:
            data[r][c] = value
        except Exception as e:
            print(e, ', Creating new column.')
            while True:
                if len(data[r]) != c:
                    data[r].append('')
                else:
                    break
            if len(data[r]) >= c:
                data[r][c-1] = value
    with open(path, 'w', newline='') as f:
        w=csv.writer(f)
        for row in data:
            w.writerow(row)

def GenerateRandomNumber(start,stop):
    return random.randrange(start,stop)